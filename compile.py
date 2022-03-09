# Libraries
import re
import os
import csv
import openpyxl
from openpyxl import load_workbook


def list_files(dir):

    #creates empty working list for files in user-defined directory
    r = []
    #goes through directory and adds all files to list
    for subdir, dirs, files in os.walk(dir):
        for filename in files:
            filepath = subdir + os.sep + filename
            r.append(filepath)
    #retuns list
    return r

def write_output(dt):

    #opens output.csv, initializes header row
    with open('output.csv','w', newline='') as result_file:
        wr = csv.writer(result_file, dialect='excel')
        wr.writerow(['respondant', 'q1', 'q2', 'q3', 'q4', 'q5', 'q6',
        'q7', 'q8', 'q9', 'q10', 'q11', 'q12', 'q13', 'q14','q15',
        'q17', 'q18', 'q19', 'q20', 'q21', 'q22', 'q23', 'aabd', 'aagp',
        'aapm', 'aait', 'aaot', 'azbd', 'azgp', 'azpm', 'azit', 'azot',
        'hlbd', 'hlgp', 'hlpm', 'hlit', 'hlot', 'nabd', 'nagp', 'napm',
        'nait', 'naot', 'pibd', 'pigp', 'pipm', 'piit', 'piot', 'whbd',
        'whgp', 'whpm', 'whit', 'whot', 'mrbd', 'mrgp', 'mrpm', 'mrit',
        'mrot', 'oebd', 'oegp', 'oepm', 'oeit', 'oeot', 'cfbd', 'cfgp',
        'cfpm', 'cfit', 'cfot', 'cmbd', 'cmgp', 'cmpm', 'cmit', 'cmot',
        'oibd', 'oigp', 'oipm', 'oiit', 'oiot', 'ogbd', 'oggp', 'ogpm',
        'ogit', 'ogot'])

    for x in dt:
        #creates empty working list
        d = []

        #finds respondant name from the file name after "-" and before ".xlsx"
        res_list = re.findall('^.*- (.*)\.xlsx$', x)
        res = ''.join(res_list)
        d.append(res)

        #hardcoded variables that point to specific cells in xlsx file
        wrkbk = load_workbook(x)
        sh = wrkbk.active
        q1 = sh.cell(row=5,column=8)
        q2 = sh.cell(row=6,column=8)
        q3 = sh.cell(row=10,column=8)
        q4 = sh.cell(row=11,column=8)
        q5 = sh.cell(row=12,column=8)
        q6 = sh.cell(row=13,column=8)
        q7 = sh.cell(row=14,column=8)
        q8 = sh.cell(row=18,column=8)
        q9 = sh.cell(row=19,column=8)
        q10 = sh.cell(row=20,column=8)
        q11 = sh.cell(row=24,column=8)
        q12 = sh.cell(row=25,column=8)
        q13 = sh.cell(row=26,column=8)
        q14 = sh.cell(row=27,column=8)
        q15 = sh.cell(row=28,column=8)
        q17 = sh.cell(row=49,column=8)
        q18 = sh.cell(row=50,column=8)
        q19 = sh.cell(row=51,column=8)
        q20 = sh.cell(row=52,column=8)
        q21 = sh.cell(row=56,column=8)
        q22 = sh.cell(row=57,column=8)
        q23 = sh.cell(row=58,column=8)
        aabd = sh.cell(row=32,column=3)
        aagp = sh.cell(row=32,column=4)
        aapm = sh.cell(row=32,column=5)
        aait = sh.cell(row=32,column=6)
        aaot = sh.cell(row=32,column=7)
        azbd = sh.cell(row=33,column=3)
        azgp = sh.cell(row=33,column=4)
        azpm = sh.cell(row=33,column=5)
        azit = sh.cell(row=33,column=6)
        azot = sh.cell(row=33,column=7)
        hlbd = sh.cell(row=34,column=3)
        hlgp = sh.cell(row=34,column=4)
        hlpm = sh.cell(row=34,column=5)
        hlit = sh.cell(row=34,column=6)
        hlot = sh.cell(row=34,column=7)
        nabd = sh.cell(row=35,column=3)
        nagp = sh.cell(row=35,column=4)
        napm = sh.cell(row=35,column=5)
        nait = sh.cell(row=35,column=6)
        naot = sh.cell(row=35,column=7)
        pibd = sh.cell(row=36,column=3)
        pigp = sh.cell(row=36,column=4)
        pipm = sh.cell(row=36,column=5)
        piit = sh.cell(row=36,column=6)
        piot = sh.cell(row=36,column=7)
        whbd = sh.cell(row=37,column=3)
        whgp = sh.cell(row=37,column=4)
        whpm = sh.cell(row=37,column=5)
        whit = sh.cell(row=37,column=6)
        whot = sh.cell(row=37,column=7)
        mrbd = sh.cell(row=38,column=3)
        mrgp = sh.cell(row=38,column=4)
        mrpm = sh.cell(row=38,column=5)
        mrit = sh.cell(row=38,column=6)
        mrot = sh.cell(row=38,column=7)
        oebd = sh.cell(row=39,column=3)
        oegp = sh.cell(row=39,column=4)
        oepm = sh.cell(row=39,column=5)
        oeit = sh.cell(row=39,column=6)
        oeot = sh.cell(row=39,column=7)
        cfbd = sh.cell(row=41,column=3)
        cfgp = sh.cell(row=41,column=4)
        cfpm = sh.cell(row=41,column=5)
        cfit = sh.cell(row=41,column=6)
        cfot = sh.cell(row=41,column=7)
        cmbd = sh.cell(row=42,column=3)
        cmgp = sh.cell(row=42,column=4)
        cmpm = sh.cell(row=42,column=5)
        cmit = sh.cell(row=42,column=6)
        cmot = sh.cell(row=42,column=7)
        oibd = sh.cell(row=43,column=3)
        oigp = sh.cell(row=43,column=4)
        oipm = sh.cell(row=43,column=5)
        oiit = sh.cell(row=43,column=6)
        oiot = sh.cell(row=43,column=7)
        ogbd = sh.cell(row=45,column=3)
        oggp = sh.cell(row=45,column=4)
        ogpm = sh.cell(row=45,column=5)
        ogit = sh.cell(row=45,column=6)
        ogot = sh.cell(row=45,column=7)

        #adds variables to list
        d.append(q1.value)
        d.append(q2.value)
        d.append(q3.value)
        d.append(q4.value)
        d.append(q5.value)
        d.append(q6.value)
        d.append(q7.value)
        d.append(q8.value)
        d.append(q9.value)
        d.append(q10.value)
        d.append(q11.value)
        d.append(q12.value)
        d.append(q13.value)
        d.append(q14.value)
        d.append(q15.value)
        d.append(q17.value)
        d.append(q18.value)
        d.append(q19.value)
        d.append(q20.value)
        d.append(q21.value)
        d.append(q22.value)
        d.append(q23.value)
        d.append(aabd.value)
        d.append(aagp.value)
        d.append(aapm.value)
        d.append(aait.value)
        d.append(aaot.value)
        d.append(azbd.value)
        d.append(azgp.value)
        d.append(azpm.value)
        d.append(azit.value)
        d.append(azot.value)
        d.append(hlbd.value)
        d.append(hlgp.value)
        d.append(hlpm.value)
        d.append(hlit.value)
        d.append(hlot.value)
        d.append(nabd.value)
        d.append(nagp.value)
        d.append(napm.value)
        d.append(nait.value)
        d.append(naot.value)
        d.append(pibd.value)
        d.append(pigp.value)
        d.append(pipm.value)
        d.append(piit.value)
        d.append(piot.value)
        d.append(whbd.value)
        d.append(whgp.value)
        d.append(whpm.value)
        d.append(whit.value)
        d.append(whot.value)
        d.append(mrbd.value)
        d.append(mrgp.value)
        d.append(mrpm.value)
        d.append(mrit.value)
        d.append(mrot.value)
        d.append(oebd.value)
        d.append(oegp.value)
        d.append(oepm.value)
        d.append(oeit.value)
        d.append(oeot.value)
        d.append(cfbd.value)
        d.append(cfgp.value)
        d.append(cfpm.value)
        d.append(cfit.value)
        d.append(cfot.value)
        d.append(cmbd.value)
        d.append(cmgp.value)
        d.append(cmpm.value)
        d.append(cmit.value)
        d.append(cmot.value)
        d.append(oibd.value)
        d.append(oigp.value)
        d.append(oipm.value)
        d.append(oiit.value)
        d.append(oiot.value)
        d.append(ogbd.value)
        d.append(oggp.value)
        d.append(ogpm.value)
        d.append(ogit.value)
        d.append(ogot.value)

        #print outputs to terminal for verification
        print (d)
        #writes list to excel in a new row
        with open('output.csv','a', newline='') as result_file:
            wr = csv.writer(result_file, dialect='excel')
            wr.writerow(d)

# user input filepath that contains surveys
x = input("Enter filepath:")

#creates working list of files in directory
y = list_files(x)

#writes values to output.csv
b = write_output(y)
